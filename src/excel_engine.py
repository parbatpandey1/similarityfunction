"""
📊 Stage 6: Excel Dashboard v6.0
✅ 11 sheets including email ready data with matched skill pair
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from utils import *
from config import *


def format_excel_header(worksheet, color='4472C4'):
    """Format Excel header"""
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    font = Font(bold=True, color='FFFFFF')

    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for column in worksheet.columns:
        max_len    = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        worksheet.column_dimensions[col_letter].width = min(max_len + 2, 50)


def build_email_ready_df(rich_df):
    """
    Build focused email-ready DataFrame.
    Includes exact matched skill pair + full profiles.
    Direct input for email_generator.py.
    """
    email_cols = [
        # Match metadata
        'final_similarity_score', 'match_quality',
        'skill_score', 'aspiration_score',

        # ✅ Exact matched skill pair
        'mentee_skill_area',    'mentee_skill_text',
        'mentor_skill_area',    'mentor_skill_text',

        # Mentee
        'mentee_name', 'mentee_email',
        'mentee_department', 'mentee_stream',
        'mentee_main_interest',       'mentee_main_level',
        'mentee_additional_interest', 'mentee_additional_level',
        'mentee_aspirations',

        # Mentor
        'mentor_name', 'mentor_email',
        'mentor_stream',
        'mentor_main_expertise',       'mentor_main_level',
        'mentor_additional_expertise', 'mentor_additional_level',
        'mentor_experience', 'mentor_work_affiliation',
        'mentor_hours_per_week', 'mentor_contact_preference',
    ]
    cols_exist = [c for c in email_cols if c in rich_df.columns]
    return rich_df[cols_exist].sort_values(
        'final_similarity_score', ascending=False
    ).reset_index(drop=True)


def create_dashboard(dedup_df, assignments_df, recommendations_df,
                     mentor_top_df, rich_df, utilization_df):
    """Create Excel dashboard"""
    print("\n📊 Creating Excel dashboard v6.0...")

    excel_path = OUTPUT_DIR / '🏆_MENTORSHIP_DASHBOARD_v6.xlsx'
    email_df   = build_email_ready_df(rich_df)

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:

        # Sheet 1: Final Assignments
        print("  [1/11] Final assignments...")
        assignment_cols = [
            'mentee_name', 'mentee_skill_text', 'mentor_name', 'mentor_skill_text',
            'skill_score', 'aspiration_score', 'final_similarity_score',
            'expertise_gap', 'semantic_similarity'
        ]
        ac_exist = [c for c in assignment_cols if c in assignments_df.columns]
        assignments_df[ac_exist].sort_values(
            'final_similarity_score', ascending=False
        ).to_excel(writer, sheet_name='✅_FINAL_ASSIGNMENTS', index=False)

        # Sheet 2: Top 5 per Mentee
        print("  [2/11] Top 5 per mentee...")
        rec_cols = [
            'recommendation_rank', 'mentee_name', 'mentee_skill_text',
            'mentor_name', 'mentor_skill_text',
            'skill_score', 'aspiration_score', 'final_similarity_score'
        ]
        rc_exist = [c for c in rec_cols if c in recommendations_df.columns]
        recommendations_df[rc_exist].to_excel(
            writer, sheet_name='🎓_TOP5_PER_MENTEE', index=False
        )

        # Sheet 3: Top 5 per Mentor (reciprocal)
        print("  [3/11] Top 5 per mentor (reciprocal)...")
        mentor_rec_cols = [
            'recommendation_rank', 'mentor_name', 'mentor_skill_text',
            'mentee_name', 'mentee_skill_text',
            'skill_score', 'aspiration_score', 'final_similarity_score'
        ]
        mrc_exist = [c for c in mentor_rec_cols if c in mentor_top_df.columns]
        mentor_top_df[mrc_exist].to_excel(
            writer, sheet_name='👨‍🏫_TOP5_PER_MENTOR', index=False
        )

        # Sheet 4: Email Ready Data
        print("  [4/11] Email ready data...")
        email_df.to_excel(writer, sheet_name='📧_EMAIL_READY_DATA', index=False)

        # Sheet 5: Capacity
        print("  [5/11] Capacity utilization...")
        utilization_df.sort_values('assigned', ascending=False).to_excel(
            writer, sheet_name='👥_CAPACITY_STATUS', index=False
        )

        # Sheet 6: Best 100
        print("  [6/11] Best 100 matches...")
        best = dedup_df.nlargest(100, 'final_similarity_score')[ac_exist]
        best.to_excel(writer, sheet_name='🔥_BEST_100_MATCHES', index=False)

        # Sheet 7: Aspiration Analysis
        print("  [7/11] Aspiration analysis...")
        asp_df = dedup_df[dedup_df['final_similarity_score'] > 0].copy()
        asp_analysis = asp_df.groupby(
            pd.cut(
                asp_df['aspiration_score'],
                bins=[0, 0.3, 0.5, 0.7, 1.0],
                labels=['Low (0-0.3)', 'Fair (0.3-0.5)',
                        'Good (0.5-0.7)', 'Excellent (0.7-1.0)']
            )
        ).agg({
            'final_similarity_score': ['count', 'mean'],
            'skill_score':            'mean',
            'aspiration_score':       'mean'
        }).round(3).reset_index()
        asp_analysis.columns = ['Aspiration Range', 'Count',
                                'Avg Final', 'Avg Skill', 'Avg Aspiration']
        asp_analysis.to_excel(writer, sheet_name='🎯_ASPIRATION_ANALYSIS', index=False)

        # Sheet 8: Stream Analysis
        print("  [8/11] Stream analysis...")
        stream = dedup_df[dedup_df['final_similarity_score'] > 0].groupby(
            ['mentee_stream', 'mentor_stream']
        ).agg({
            'final_similarity_score': ['count', 'mean', 'max'],
            'skill_score':            'mean',
            'aspiration_score':       'mean'
        }).round(3).reset_index()
        stream.columns = ['Mentee Stream', 'Mentor Stream', 'Count',
                          'Avg Final', 'Max Final', 'Avg Skill', 'Avg Aspiration']
        stream.sort_values('Avg Final', ascending=False).to_excel(
            writer, sheet_name='🌊_STREAM_ANALYSIS', index=False
        )

        # Sheet 9: Skill Areas
        print("  [9/11] Skill area analysis...")
        skills = dedup_df[dedup_df['final_similarity_score'] > 0].groupby(
            ['mentee_skill_area', 'mentor_skill_area']
        ).agg({
            'final_similarity_score': ['count', 'mean', 'max']
        }).round(3).reset_index()
        skills.columns = ['Mentee Skill Area', 'Mentor Skill Area', 'Count', 'Avg', 'Max']
        skills.sort_values('Avg', ascending=False).to_excel(
            writer, sheet_name='🎯_SKILL_AREAS', index=False
        )

        # Sheet 10: Match Quality Breakdown
        print("  [10/11] Match quality breakdown...")
        quality_df = email_df.groupby('match_quality').agg(
            count       =('final_similarity_score', 'count'),
            avg_final   =('final_similarity_score', 'mean'),
            avg_skill   =('skill_score',            'mean'),
            avg_asp     =('aspiration_score',        'mean'),
        ).round(3).reset_index()
        quality_df.columns = ['Match Quality', 'Count',
                              'Avg Final', 'Avg Skill', 'Avg Aspiration']
        quality_order = ['Excellent', 'Good', 'Fair', 'Weak']
        quality_df['_ord'] = quality_df['Match Quality'].map(
            {q: i for i, q in enumerate(quality_order)}
        )
        quality_df.sort_values('_ord').drop('_ord', axis=1).to_excel(
            writer, sheet_name='📊_MATCH_QUALITY', index=False
        )

        # Sheet 11: Summary
        print("  [11/11] Executive summary...")
        total_mentees = dedup_df['mentee_person_id'].nunique()
        assigned      = len(assignments_df)

        summary = pd.DataFrame([
            {'Metric': '📊 SYSTEM VERSION',               'Value': 'VectorBridge v6.0'},
            {'Metric': '',                                 'Value': ''},
            {'Metric': 'Total Mentees',                   'Value': total_mentees},
            {'Metric': 'Assigned Mentees',                'Value': assigned},
            {'Metric': 'Unassigned',                      'Value': total_mentees - assigned},
            {'Metric': 'Assignment Rate',                 'Value': f"{100*assigned/total_mentees:.1f}%"},
            {'Metric': '',                                 'Value': ''},
            {'Metric': 'Total Mentors',                   'Value': dedup_df['mentor_person_id'].nunique()},
            {'Metric': 'Active Mentors',                  'Value': int((utilization_df['assigned'] > 0).sum())},
            {'Metric': 'Total Capacity',                  'Value': int(utilization_df['capacity'].sum())},
            {'Metric': 'Capacity Used',                   'Value': int(utilization_df['assigned'].sum())},
            {'Metric': '',                                 'Value': ''},
            {'Metric': 'Best Final Score',                'Value': f"{dedup_df['final_similarity_score'].max():.4f}"},
            {'Metric': 'Avg Final Score (assigned)',      'Value': f"{assignments_df['final_similarity_score'].mean():.4f}"},
            {'Metric': 'Avg Skill Score (assigned)',      'Value': f"{assignments_df['skill_score'].mean():.4f}"},
            {'Metric': 'Avg Aspiration Score (assigned)', 'Value': f"{assignments_df['aspiration_score'].mean():.4f}"},
            {'Metric': '',                                 'Value': ''},
            {'Metric': 'Excellent Matches (>=0.75)',      'Value': int((email_df['match_quality'] == 'Excellent').sum())},
            {'Metric': 'Good Matches (>=0.55)',           'Value': int((email_df['match_quality'] == 'Good').sum())},
            {'Metric': 'Fair Matches (>=0.40)',           'Value': int((email_df['match_quality'] == 'Fair').sum())},
            {'Metric': 'Weak Matches (<0.40)',            'Value': int((email_df['match_quality'] == 'Weak').sum())},
            {'Metric': '',                                 'Value': ''},
            {'Metric': '🎯 FORMULA',                      'Value': f'Final = {SKILL_WEIGHT:.0%} Skill + {ASPIRATION_WEIGHT:.0%} Aspiration'},
            {'Metric': '📧 EMAIL INPUT',                  'Value': 'email_ready_data → email_generator.py'},
        ])
        summary.to_excel(writer, sheet_name='📈_SUMMARY', index=False)

    # Apply formatting
    print("  🎨 Formatting...")
    workbook = load_workbook(excel_path)
    sheet_colors = {
        '✅_FINAL_ASSIGNMENTS':   '217346',
        '🎓_TOP5_PER_MENTEE':     '4472C4',
        '👨‍🏫_TOP5_PER_MENTOR':    '7030A0',
        '📧_EMAIL_READY_DATA':    'C55A11',
        '👥_CAPACITY_STATUS':     '2F75B6',
        '🔥_BEST_100_MATCHES':    '404040',
        '🎯_ASPIRATION_ANALYSIS': '843C0C',
        '🌊_STREAM_ANALYSIS':     '1F4E79',
        '🎯_SKILL_AREAS':         '375623',
        '📊_MATCH_QUALITY':       '7B2C2C',
        '📈_SUMMARY':             '4472C4',
    }
    for sheet_name in workbook.sheetnames:
        color = sheet_colors.get(sheet_name, '4472C4')
        format_excel_header(workbook[sheet_name], color)
    workbook.save(excel_path)

    print(f"  ✅ Dashboard: {excel_path.name}")


def main():
    """Main Excel engine"""
    print("\n" + "="*80)
    print("STAGE 6: EXCEL DASHBOARD v6.0")
    print("="*80)

    try:
        print("\n[1/2] Loading data...")
        dedup_df           = load_from_output('deduplicated_matches.csv')
        assignments_df     = load_from_output('final_assignments.csv')
        recommendations_df = load_from_output('top_n_recommendations.csv')
        mentor_top_df      = load_from_output('top_n_per_mentor.csv')
        rich_df            = load_from_output('rich_matched_dataset.csv')
        utilization_df     = load_from_output('mentor_utilization.csv')

        print("\n[2/2] Creating dashboard...")
        create_dashboard(dedup_df, assignments_df, recommendations_df,
                         mentor_top_df, rich_df, utilization_df)

        print("\n📊 DASHBOARD COMPLETE")
        print("  11 sheets created:")
        print("  1.  ✅_FINAL_ASSIGNMENTS")
        print("  2.  🎓_TOP5_PER_MENTEE")
        print("  3.  👨‍🏫_TOP5_PER_MENTOR         ← reciprocal")
        print("  4.  📧_EMAIL_READY_DATA         ← matched pair + profiles ✨")
        print("  5.  👥_CAPACITY_STATUS")
        print("  6.  🔥_BEST_100_MATCHES")
        print("  7.  🎯_ASPIRATION_ANALYSIS")
        print("  8.  🌊_STREAM_ANALYSIS")
        print("  9.  🎯_SKILL_AREAS")
        print("  10. 📊_MATCH_QUALITY")
        print("  11. 📈_SUMMARY")

        print("\n✅ EXCEL ENGINE COMPLETE")

    except Exception as e:
        print(f"\n❌ FAILED: {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    main()
